import os
import random
from copy import copy
from datetime import datetime
from typing import Dict, Tuple

import excel2img
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, PatternFill
from PIL import Image
from sklearn.preprocessing import MinMaxScaler

from com.statlume.nba.database import Database


def get_threepm_score(row: Dict) -> float:
    mpg_season = row["MPG"]
    mpg_last3 = row["Last 3 MPG"]
    attempt_rate = row["3PAr"]
    values = {
        "3pmrank": row["Vs 3PM Rank"],
        "3pprank": row["Vs 3P% Rank"],
        "rate": attempt_rate,
        "mpg": mpg_last3 - mpg_season,
    }
    weights = {
        "3pmrank": 0.05,
        "3pprank": 0.05,
        "rate": 0.60,
        "mpg": 0.30,
    }
    score = sum(float(values[col]) * weights[col] for col in weights)
    return score


def get_line_odds(player: str) -> Tuple[str, str]:
    odds = Database('nba').select_player_odds(player=player, field="player_threes")
    return (odds[0], odds[1])


def get_team_name(team_abv: str) -> str:
    teams = Database("nba").select_table_as_df("teams")
    for index, team in teams.iterrows():
        if team["abbreviation"] == team_abv:
            return team["full_name"]


def get_opponent_team(team: str) -> str:
    schedule = Database("nba").select_table_as_df("schedule")
    today = pd.to_datetime(datetime.today().date())
    todays_games = schedule[schedule["GameDate"] == today]
    for index, game in todays_games.iterrows():
        if game["Home"] == team:
            return game["Visitor"]
        if game["Visitor"] == team:
            return game["Home"]


def get_opponent_rank(opponent: str, stat: str) -> int:
    teamstats = Database("nba").select_table_as_df("teamstats")
    for index, team in teamstats.iterrows():
        if team["Team"] == opponent:
            return round(team[stat + "_RANK"])


def get_player_id(player_name: str) -> int:
    players = Database("nba").select_table_as_df("players")
    for index, player in players.iterrows():
        if player["full_name"] == player_name:
            return player["Player_ID"]


def get_last_games_stat(player_id: int, stat: str, games: int) -> float:
    gamelogs = Database("nba").select_table_as_df("gamelogs")
    gamelogs = gamelogs[gamelogs["Player_ID"] == player_id]
    gamelogs = gamelogs.sort_values(by="GAME_DATE", ascending=False)
    return round(gamelogs.head(games)[stat].mean(), 1)


def get_season_stat(player_id: int, stat: str, team: str) -> float:
    player_stats = Database("nba").select_table_as_df("playerstats")
    for index, player in player_stats.iterrows():
        if player["Player_ID"] == str(player_id):
            if player["Team"] == team:
                return player[stat]


def filter_injuries(threes: pd.DataFrame) -> pd.DataFrame:
    injuries = Database("nba").select_table_as_df("injuries")
    for index, player in threes.iterrows():
        for injury_index, injury in injuries.iterrows():
            if player["Player"] == injury["Player"]:
                if injury["Status"] == "O":
                    threes = threes.drop(index)
                if injury["Status"] == "DTD":
                    threes.loc[index, "Player"] = player["Player"] + "*"
    return threes


def get_threes() -> pd.DataFrame:
    lineups = Database("nba").select_table_as_df("lineups")
    threes = pd.DataFrame()
    for index, player in lineups.iterrows():
        player_name: str = player["Player"]
        id = get_player_id(player_name=player_name)
        team_abv = player["Abbreviation"]
        team = get_team_name(team_abv=team_abv)
        opponent = get_opponent_team(team=team)
        three_rank = get_opponent_rank(opponent=opponent, stat="3P")
        percent_rank = get_opponent_rank(opponent=opponent, stat="3PP")
        season_stat = get_season_stat(player_id=id, stat="FG3M", team=team)
        minutes_season = get_season_stat(player_id=id, stat="MP", team=team)
        attempt_rate = get_season_stat(player_id=id, stat="3PR", team=team)
        minutes_last = get_last_games_stat(player_id=id, stat="MIN", games=3)
        recent_stat = get_last_games_stat(player_id=id, stat="FG3M", games=3)
        try:
            line, odds = get_line_odds(player=player_name)
        except:
            continue
        values = {
            "Team": team_abv,
            "Player": player_name,
            "Opponent": opponent,
            "Vs 3PM Rank": three_rank,
            "Vs 3P% Rank": percent_rank,
            "MPG": minutes_season,
            "Last 3 MPG": minutes_last,
            "Season Avg": season_stat,
            "Last 3 Avg": recent_stat,
            "3PAr": attempt_rate,
            "Line": line,
            "Odds": odds,
        }
        df = pd.DataFrame([values]).dropna()
        if df.empty:
            continue
        df["Score"] = get_threepm_score(values)
        threes = pd.concat([threes, df], ignore_index=True)
        print(player_name)
    minmax = MinMaxScaler(feature_range=(10.01, 89.99))
    threes["Score"] = minmax.fit_transform(
        threes["Score"].values.reshape(-1, 1)
    ) - random.uniform(1.01, 1.99)
    threes = threes.round(2)
    return threes.sort_values(by="Score", ascending=False)


def get_threes_sheet(threes: pd.DataFrame) -> None:
    wb = Workbook()
    ws = wb.active

    columns = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]
    index = 0
    row = 6
    color = PatternFill(start_color="221E2F", fill_type="solid")
    font = Font(color="ffffff", bold=True)
    for col in threes.columns[0:]:
        ws[columns[index] + str(row)] = col
        ws[columns[index] + str(row)].fill = color
        ws[columns[index] + str(row)].font = font
        index += 1
    row = 7
    for index, mismatch in threes.head(20).iterrows():
        ws[columns[0] + str(row)] = mismatch["Team"]
        ws[columns[1] + str(row)] = mismatch["Player"]
        ws[columns[2] + str(row)] = mismatch["Opponent"]
        ws[columns[3] + str(row)] = mismatch["Vs 3PM Rank"]
        ws[columns[4] + str(row)] = mismatch["Vs 3P% Rank"]
        ws[columns[5] + str(row)] = mismatch["MPG"]
        ws[columns[6] + str(row)] = mismatch["Last 3 MPG"]
        ws[columns[7] + str(row)] = mismatch["Season Avg"]
        ws[columns[8] + str(row)] = mismatch["Last 3 Avg"]
        ws[columns[9] + str(row)] = mismatch["3PAr"]
        ws[columns[10] + str(row)] = mismatch["Line"]
        ws[columns[11] + str(row)] = mismatch["Odds"]
        ws[columns[12] + str(row)] = mismatch["Score"]
        if row % 2 == 0:
            color = PatternFill(start_color="d3d3d3", fill_type="solid")
        else:
            color = PatternFill(start_color="ffffff", fill_type="solid")
        for col in columns:
            ws[col + str(row)].fill = color
        row += 1
    ws.merge_cells(columns[0] + str(row) + ":" + columns[-1] + str(row))
    ws[
        columns[0] + str(row)
    ] = """
        * Players with asterisks have been listed as Day To Day for injuries
    """
    for col in columns:
        ws.column_dimensions[col].width = 12
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25
    ws.merge_cells("C1:E5")
    ws["C1"] = "NBA Three Pointers"
    ws["C1"].font = Font(size="24")
    for col in ws.columns:
        for cell in col:
            alignment_obj = copy(cell.alignment)
            alignment_obj.horizontal = "center"
            alignment_obj.vertical = "center"
            cell.alignment = alignment_obj
            bold_obj = copy(cell.font)
            bold_obj.bold = True
            cell.font = bold_obj
    alignment_obj = copy(ws["C1"].alignment)
    alignment_obj.horizontal = "right"
    ws["C1"].alignment = alignment_obj
    ws.merge_cells("F1:H5")
    pil_image = Image.open("assets/logo-transparent.png")
    pil_image = pil_image.resize((280, 100))
    temp_image_path = "assets/temp1.png"
    pil_image.save(temp_image_path)
    img = XLImage(temp_image_path)
    ws.add_image(img, "F1")
    color = PatternFill(start_color="ffffff", fill_type="solid")
    for col in columns:
        for cell in range(1, 6):
            ws[col + str(cell)].fill = color
    rule = ColorScaleRule(
        start_type="min",
        start_value=50,
        start_color="b7f6b1",
        end_type="max",
        end_value=100,
        end_color="35e524",
    )
    ws.conditional_formatting.add("M7:M27", rule)
    path = "./com/statlume/nba/sheets/threes.xlsx"
    wb.save(path)
    os.remove("assets/temp1.png")
    image_file = "./com/statlume/nba/images/threes.png"
    excel2img.export_img(path, image_file, "Sheet", None)


if __name__ == "__main__":
    get_threes_sheet(filter_injuries(get_threes()))
