import os
import random
from copy import copy
from datetime import datetime
from typing import Any, Dict, Tuple

import excel2img
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image
from sklearn.preprocessing import MinMaxScaler

from com.statlume.nba.database import Database


def get_mismatch_score(row: Dict[Any, Any]) -> float:
    avg = row["Season Avg"]
    last = row["Last 5 Avg"]
    line = row["Line"]
    mpg_season = row["MPG"]
    mpg_last3 = row["Last 3 MPG"]
    values = {
        "rank": row["Vs Rank"],
        "mpg": mpg_last3 - mpg_season,
        "performance": last - avg,
        "line": avg - line,
    }
    weights = {"rank": 0.25, "mpg": 0.30, "performance": 0.15, "line": 0.30}
    score = sum(float(values[col]) * weights[col] for col in weights)
    return score


def get_line_odds(player: str, field: str) -> Tuple[str, str]:
    if field == "PTS":
        field = "player_points"
    if field == "AST":
        field = "player_assists"
    if field == "REB":
        field = "player_rebounds"
    odds = Database('nba').select_player_odds(player=player, field=field)
    return (odds[0], odds[1])


def get_team_name(team_abv: str) -> str:
    teams = Database("nba").select_table_as_df("teams")
    for index, team in teams.iterrows():
        if team["abbreviation"] == team_abv:
            return team["full_name"]


def get_opponent_team(team: str) -> str:
    schedule = Database("nba").select_table_as_df("schedule")
    today = pd.to_datetime(datetime.today().date())
    todays_games = schedule[schedule["GameDate"] == today]
    for index, game in todays_games.iterrows():
        if game["Home"] == team:
            return game["Visitor"]
        if game["Visitor"] == team:
            return game["Home"]


def get_opponent_rank(opponent: str, stat: str) -> int:
    teamstats = Database("nba").select_table_as_df("teamstats")
    for index, team in teamstats.iterrows():
        if team["Team"] == opponent:
            return round(team[stat + "_RANK"])


def get_player_id(player_name: str) -> int:
    players = Database("nba").select_table_as_df("players")
    for index, player in players.iterrows():
        if player["full_name"] == player_name:
            return player["Player_ID"]


def get_last_games_stat(player_id: int, stat: str, games: int) -> float:
    gamelogs = Database("nba").select_table_as_df("gamelogs")
    gamelogs = gamelogs[gamelogs["Player_ID"] == player_id]
    gamelogs = gamelogs.sort_values(by="GAME_DATE", ascending=False)
    return round(gamelogs.head(games)[stat].mean(), 1)


def get_season_stat(player_id: int, stat: str, team: str) -> float:
    player_stats = Database("nba").select_table_as_df("playerstats")
    for index, player in player_stats.iterrows():
        if player["Player_ID"] == str(player_id):
            if player["Team"] == team:
                return player[stat]


def filter_injuries(mismatches: pd.DataFrame) -> pd.DataFrame:
    injuries = Database("nba").select_table_as_df("injuries")
    for index, player in mismatches.iterrows():
        for injury_index, injury in injuries.iterrows():
            if player["Player"] == injury["Player"]:
                if injury["Status"] == "O":
                    mismatches = mismatches.drop(index)
                if injury["Status"] == "DTD":
                    mismatches.loc[index, "Player"] = player["Player"] + "*"
    return mismatches


def delete_matchups():
    images = "./com/statlume/nba/images/"
    for filename in os.listdir(images):
        file_path = os.path.join(images, filename)
        if os.path.isfile(file_path):
            os.remove(file_path)

    sheets = "./com/statlume/nba/sheets/"
    for filename in os.listdir(sheets):
        file_path = os.path.join(sheets, filename)
        if os.path.isfile(file_path):
            os.remove(file_path)


def get_mismatches() -> pd.DataFrame:
    lineups = Database("nba").select_table_as_df("lineups")
    mismatches = pd.DataFrame()
    for index, player in lineups.iterrows():
        player_name = player["Player"]
        id = get_player_id(player_name=player_name)
        team_abv = player["Abbreviation"]
        team = get_team_name(team_abv=team_abv)
        opponent = get_opponent_team(team=team)
        stats = ["PTS", "REB", "AST"]
        for stat in stats:
            opponent_rank = get_opponent_rank(opponent=opponent, stat=stat)
            season_stat = get_season_stat(player_id=id, stat=stat, team=team)
            recent_stat = get_last_games_stat(player_id=id, stat=stat, games=5)
            minutes_season = get_season_stat(player_id=id, stat="MP", team=team)
            minutes_last = get_last_games_stat(player_id=id, stat="MIN", games=3)
            try:
                line, odds = get_line_odds(player=player_name, field=stat)
            except:
                continue
            values = {
                "Team": team_abv,
                "Player": player_name,
                "Stat": stat,
                "Opponent": opponent,
                "Vs Rank": opponent_rank,
                "MPG": minutes_season,
                "Last 3 MPG": minutes_last,
                "Season Avg": season_stat,
                "Last 5 Avg": recent_stat,
                "Line": line,
                "Odds": odds,
            }
            df = pd.DataFrame([values]).dropna()
            if df.empty:
                continue
            df["Score"] = get_mismatch_score(values)
            mismatches = pd.concat([mismatches, df], ignore_index=True)
        print(player_name)
    minmax = MinMaxScaler(feature_range=(10.01, 89.99))
    mismatches["Score"] = minmax.fit_transform(
        mismatches["Score"].values.reshape(-1, 1)
    ) - random.uniform(1.01, 1.99)
    mismatches = mismatches.round(2)
    return mismatches.sort_values(by="Score", ascending=False)


def get_mismatch_sheet(mismatches: pd.DataFrame) -> None:
    wb = Workbook()
    ws = wb.active

    columns = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]
    index = 0
    row = 6
    color = PatternFill(start_color="221E2F", fill_type="solid")
    font = Font(color="ffffff", bold=True)
    for col in mismatches.columns[0::]:
        ws[columns[index] + str(row)] = col
        ws[columns[index] + str(row)].fill = color
        ws[columns[index] + str(row)].font = font
        index += 1
    row = 7
    for index, mismatch in mismatches.head(20).iterrows():
        ws[columns[0] + str(row)] = mismatch["Team"]
        ws[columns[1] + str(row)] = mismatch["Player"]
        ws[columns[2] + str(row)] = mismatch["Stat"]
        ws[columns[3] + str(row)] = mismatch["Opponent"]
        ws[columns[4] + str(row)] = mismatch["Vs Rank"]
        ws[columns[5] + str(row)] = mismatch["MPG"]
        ws[columns[6] + str(row)] = mismatch["Last 3 MPG"]
        ws[columns[7] + str(row)] = mismatch["Season Avg"]
        ws[columns[8] + str(row)] = mismatch["Last 5 Avg"]
        ws[columns[9] + str(row)] = mismatch["Line"]
        ws[columns[10] + str(row)] = mismatch["Odds"]
        ws[columns[11] + str(row)] = mismatch["Score"]
        if row % 2 == 0:
            color = PatternFill(start_color="d3d3d3", fill_type="solid")
        else:
            color = PatternFill(start_color="ffffff", fill_type="solid")
        for col in columns:
            ws[col + str(row)].fill = color
        row += 1
    ws.merge_cells(columns[0] + str(row) + ":" + columns[-1] + str(row))
    ws[
        columns[0] + str(row)
    ] = """
        * Players with asterisks have been listed as Day To Day for injuries
    """
    for col in columns:
        ws.column_dimensions[col].width = 12
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["D"].width = 25
    ws.merge_cells("D1:E5")
    ws["D1"] = " NBA Mismatches (Overs)"
    ws["D1"].font = Font(size="24")
    ws["D1"].alignment = Alignment(wrap_text=True)
    for col in ws.columns:
        for cell in col:
            alignment_obj = copy(cell.alignment)
            alignment_obj.horizontal = "center"
            alignment_obj.vertical = "center"
            cell.alignment = alignment_obj
            bold_obj = copy(cell.font)
            bold_obj.bold = True
            cell.font = bold_obj
    ws.merge_cells("F1:H5")
    pil_image = Image.open("assets/logo-transparent.png")
    pil_image = pil_image.resize((280, 100))
    temp_image_path = "assets/temp1.png"
    pil_image.save(temp_image_path)
    img = XLImage(temp_image_path)
    ws.add_image(img, "F1")
    color = PatternFill(start_color="ffffff", fill_type="solid")
    for col in columns:
        for cell in range(1, 6):
            ws[col + str(cell)].fill = color
    rule = ColorScaleRule(
        start_type="min",
        start_value=50,
        start_color="b7f6b1",
        end_type="max",
        end_value=100,
        end_color="35e524",
    )
    ws.conditional_formatting.add("L7:L27", rule)
    path = "./com/statlume/nba/sheets/mismatch.xlsx"
    wb.save(path)
    os.remove("assets/temp1.png")
    image_file = "./com/statlume/nba/images/mismatch.png"
    excel2img.export_img(path, image_file, "Sheet", None)


if __name__ == "__main__":
    get_mismatch_sheet(filter_injuries(get_mismatches()))
