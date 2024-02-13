import datetime
import os
from copy import copy

import excel2img
import pandas as pd
from database import Database
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, GradientFill, PatternFill, colors
from PIL import Image

from com.statlume.odds.odds import OddsAPI

color_codes = {
    "Atlanta Hawks": "e03a3e",
    "Boston Celtics": "008248",
    "Brooklyn Nets": "000000",
    "Charlotte Hornets": "1d1160",
    "Chicago Bulls": "ce1141",
    "Cleveland Cavaliers": "6f2633",
    "Dallas Mavericks": "007dc5",
    "Denver Nuggets": "0d2240",
    "Detroit Pistons": "c8102e",
    "Golden State Warriors": "fdb927",
    "Houston Rockets": "ce1141",
    "Indiana Pacers": "fdbb30",
    "Los Angeles Clippers": "c8102e",
    "Los Angeles Lakers": "552583",
    "Memphis Grizzlies": "5d76a9",
    "Miami Heat": "98002e",
    "Milwaukee Bucks": "00471b",
    "Minnesota Timberwolves": "0c2340",
    "New Orleans Pelicans": "002b5c",
    "New York Knicks": "f58426",
    "Oklahoma City Thunder": "007ac1",
    "Orlando Magic": "0b77bd",
    "Philadelphia 76ers": "ed174c",
    "Phoenix Suns": "1d1160",
    "Portland Trail Blazers": "e03a3e",
    "Sacramento Kings": "5b2b82",
    "San Antonio Spurs": "000000",
    "Toronto Raptors": "cd1141",
    "Utah Jazz": "0d2240",
    "Washington Wizards": "002b5c",
}


def create_team_sheet(date: str, home: pd.DataFrame, visitor: pd.DataFrame):
    wb = Workbook()
    ws = wb.active
    visitor_team_name: str = visitor["Team"].values[0]
    home_team_name: str = home["Team"].values[0]
    ws["A7"] = visitor_team_name
    ws["E7"] = visitor_team_name
    ws["C7"] = home_team_name
    ws["G7"] = home_team_name

    columns = [
        "FGPercent",
        "3PPercent",
        "2PPercent",
        "FTPercent",
        "OffReboundsPerGame",
        "DefReboundsPerGame",
        "ReboundsPerGame",
        "AssistsPerGame",
        "StealsPerGame",
        "BlocksPerGame",
        "PointsPerGame",
        "StrengthOfSchedule",
        "OffRating",
        "EffectiveFG",
        "FTPFG",
        "OppTurnoversPerGame",
        "OppTurnoverPercent",
        "TurnoverPercent",
        "TurnoversPerGame",
        "DefRating",
        "OppFGPercent",
        "Opp3PPercent",
        "Opp2PPercent",
        "OppFTPercent",
        "OppOffReboundsPerGame",
        "OppDefReboundsPerGame",
        "OppReboundsPerGame",
        "OppAssistsPerGame",
        "OppStealsPerGame",
        "OppBlocksPerGame",
        "OppPointsPerGame",
        "OppEffectiveFG",
        "OppDefReboundPercent",
        "OppFTPFG",
    ]

    for col in ws.columns:
        for cell in col:
            cell.fill = PatternFill(start_color="ffffff", fill_type="solid")

    green = "b1ffb1"
    visitor_color = color_codes[visitor_team_name]
    home_color = color_codes[home_team_name]

    row = 8
    col_a = "A"
    col_b = "B"
    col_c = "C"
    col_d = "D"
    longest_word = ""
    for col in columns[0:17]:
        if len(col) > len(longest_word):
            longest_word = col
        visitor_cell = col_a + str(row)
        stat_cell = col_b + str(row)
        home_cell = col_c + str(row)
        space_cell = col_d + str(row)
        ws[space_cell].fill = PatternFill(start_color="ffffff", fill_type="solid")

        ws[home_cell] = home[col].values[0]
        ws[stat_cell] = col
        ws[visitor_cell] = visitor[col].values[0]

        ws[home_cell].fill = PatternFill(start_color="ffffff", fill_type="solid")
        ws[stat_cell].fill = PatternFill(start_color="ffffff", fill_type="solid")
        ws[visitor_cell].fill = PatternFill(start_color="ffffff", fill_type="solid")

        if home[col].values[0] > visitor[col].values[0]:
            ws[home_cell].fill = PatternFill(start_color=home_color, fill_type="solid")
            ws[home_cell].font = Font(color=colors.WHITE)
        if home[col].values[0] < visitor[col].values[0]:
            ws[visitor_cell].fill = PatternFill(
                start_color=visitor_color, fill_type="solid"
            )
            ws[visitor_cell].font = Font(color=colors.WHITE)

        row += 1

    row = 8
    col_e = "E"
    col_f = "F"
    col_g = "G"
    longest_word = ""
    for col in columns[17:34]:
        if len(col) > len(longest_word):
            longest_word = col
        visitor_cell = col_e + str(row)
        stat_cell = col_f + str(row)
        home_cell = col_g + str(row)
        space_cell = col_d + str(row)
        ws[space_cell].fill = PatternFill(start_color="ffffff", fill_type="solid")
        ws[home_cell] = home[col].values[0]
        ws[stat_cell] = col
        ws[visitor_cell] = visitor[col].values[0]

        ws[home_cell].fill = PatternFill(start_color="ffffff", fill_type="solid")
        ws[stat_cell].fill = PatternFill(start_color="ffffff", fill_type="solid")
        ws[visitor_cell].fill = PatternFill(start_color="ffffff", fill_type="solid")

        if home[col].values[0] < visitor[col].values[0]:
            ws[home_cell].fill = PatternFill(start_color=home_color, fill_type="solid")
            ws[home_cell].font = Font(color=colors.WHITE)
        if home[col].values[0] > visitor[col].values[0]:
            ws[visitor_cell].fill = PatternFill(
                start_color=visitor_color, fill_type="solid"
            )
            ws[visitor_cell].font = Font(color=colors.WHITE)

        row += 1

    ws.merge_cells("B1:B6")
    img = Image.open("assets/logo-transparent-vertical.png")
    img_w, img_h = img.size
    background = Image.new("RGBA", (665, 475), (255, 255, 255, 255))
    bg_w, bg_h = background.size
    offset = ((bg_w - img_w) // 2, (bg_h - img_h) // 2)
    background.paste(img, offset)
    background.thumbnail((175, 120), Image.LANCZOS)
    temp_image_path = "assets/temp1.png"
    background.save(temp_image_path)
    img = XLImage(temp_image_path)
    ws.add_image(img, "B1")
    ws.merge_cells("F1:F6")
    img = XLImage(temp_image_path)
    ws.add_image(img, "F1")

    ws.merge_cells("A1:A6")
    img = Image.open(
        "assets/logos/" + visitor_team_name.lower().replace(" ", "_") + ".png"
    )
    img.thumbnail((1400, 960), Image.LANCZOS)
    img_w, img_h = img.size
    background = Image.new("RGBA", (1750, 1200), (255, 255, 255, 255))
    bg_w, bg_h = background.size
    offset = ((bg_w - img_w) // 2, (bg_h - img_h) // 2)
    background.paste(img, offset)
    background.thumbnail((175, 120), Image.LANCZOS)
    temp_image_path = "assets/temp2.png"
    background.save(temp_image_path)
    img = XLImage(temp_image_path)
    ws.add_image(img, "A1")
    ws.merge_cells("E1:E6")
    img = XLImage(temp_image_path)
    ws.add_image(img, "E1")

    ws.merge_cells("C1:C6")
    img = Image.open(
        "./assets/logos/" + home_team_name.lower().replace(" ", "_") + ".png"
    )
    img.thumbnail((1400, 960), Image.LANCZOS)
    img_w, img_h = img.size
    background = Image.new("RGBA", (1750, 1200), (255, 255, 255, 255))
    bg_w, bg_h = background.size
    offset = ((bg_w - img_w) // 2, (bg_h - img_h) // 2)
    background.paste(img, offset)
    background.thumbnail((175, 120), Image.LANCZOS)
    temp_image_path = "./assets/temp3.png"
    background.save(temp_image_path)
    img = XLImage(temp_image_path)
    ws.add_image(img, "C1")
    ws.merge_cells("G1:G6")
    img = XLImage(temp_image_path)
    ws.add_image(img, "G1")

    ws["B7"] = "@"
    ws["F7"] = "@"

    ws["A7"].fill = PatternFill(start_color=visitor_color, fill_type="solid")
    ws["A7"].font = Font(color=colors.WHITE)
    ws["E7"].fill = PatternFill(start_color=visitor_color, fill_type="solid")
    ws["E7"].font = Font(color=colors.WHITE)

    ws["C7"].fill = PatternFill(start_color=home_color, fill_type="solid")
    ws["C7"].font = Font(color=colors.WHITE)
    ws["G7"].fill = PatternFill(start_color=home_color, fill_type="solid")
    ws["G7"].font = Font(color=colors.WHITE)

    ws["B7"].fill = GradientFill(stop=(visitor_color, home_color))
    ws["F7"].fill = GradientFill(stop=(visitor_color, home_color))
    ws["B7"].font = Font(color=colors.WHITE)
    ws["F7"].font = Font(color=colors.WHITE)

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 5
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["G"].width = 25

    for col in ws.columns:
        for cell in col:
            alignment_obj = copy(cell.alignment)
            alignment_obj.horizontal = "center"
            alignment_obj.vertical = "center"
            cell.alignment = alignment_obj
            bold_obj = copy(cell.font)
            bold_obj.bold = True
            cell.font = bold_obj

    home_abv = home["Abbreviation"].values[0]
    visitor_abv = visitor["Abbreviation"].values[0]
    file_name = str(visitor_abv) + "@" + str(home_abv) + "-teams-" + date
    path = "./com/statlume/nba/sheets/" + file_name + ".xlsx"
    wb.save(path)
    os.remove("assets/temp1.png")
    os.remove("assets/temp2.png")
    os.remove("assets/temp3.png")
    load_image(path=path, file_name=file_name)


def get_recent_stats(player_id: str, logs: pd.DataFrame, field: str):
    player_logs = logs[logs["Player_ID"] == player_id]
    recent_three = player_logs.sort_values(by="GAME_DATE", ascending=False).head(3)
    average_value = recent_three[field].mean()
    return round(average_value, 1)


def create_player_sheet(
    date: str,
    home: pd.DataFrame,
    visitor: pd.DataFrame,
    home_player_stats: pd.DataFrame,
    visitor_player_stats: pd.DataFrame,
    home_player_logs: pd.DataFrame,
    visitor_player_logs: pd.DataFrame,
):
    wb = Workbook()
    ws = wb.active

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 20
    ws.column_dimensions["I"].width = 20
    ws.column_dimensions["J"].width = 20

    ws.merge_cells("C1:E1")
    visitor_team_name: str = str(visitor["Team"].values[0])
    ws["C1"] = visitor_team_name
    visitor_color = color_codes[visitor["Team"].values[0]]
    ws["C1"].fill = PatternFill(start_color=visitor_color, fill_type="solid")
    ws["C1"].font = Font(color=colors.WHITE)

    ws.merge_cells("F1:H1")
    home_team_name: str = str(home["Team"].values[0])
    ws["F1"] = home_team_name
    home_color = color_codes[home["Team"].values[0]]
    ws["F1"].fill = PatternFill(start_color=home_color, fill_type="solid")
    ws["F1"].font = Font(color=colors.WHITE)

    ws["F1"].fill = PatternFill(start_color=home_color, fill_type="solid")

    team_columns = [
        ("PointsPerGame", "Points"),
        ("ReboundsPerGame", "Rebounds"),
        ("AssistsPerGame", "Assists"),
        ("StealsPerGame", "Steals"),
        ("BlocksPerGame", "Blocks"),
    ]
    player_columns = [
        ("PTS", "Points", "player_points"),
        ("FG3M", "Three Pointers Made", "player_threes"),
        ("REB", "Rebounds", "player_rebounds"),
        ("AST", "Assists", "player_assists"),
        ("STL", "Steals", "player_steals"),
        ("BLK", "Blocks", "player_blocks"),
    ]

    row = 2
    col_a = "A"
    col_b = "B"
    col_c = "C"
    col_d = "D"
    col_e = "E"
    col_f = "F"
    col_g = "G"
    col_h = "H"
    col_i = "I"
    col_j = "J"


    injuries = Database('nba').select_table_as_df('injuries')

    for col in team_columns:
        visitor_cell_tag = col_c + str(row)
        visitor_cell_stat = col_d + str(row)
        home_cell_stat = col_g + str(row)
        home_cell_tag = col_h + str(row)
        ws[visitor_cell_tag] = col[1]
        ws[visitor_cell_stat] = visitor[col[0]].values[0]
        ws[home_cell_stat] = home[col[0]].values[0]
        ws[home_cell_tag] = col[1]
        row += 1

    for col in player_columns:
        visitor_category_start = col_a + str(row)
        visitor_category_end = col_e + str(row)
        home_category_start = col_f + str(row)
        home_category_end = col_j + str(row)
        ws.merge_cells(visitor_category_start + ":" + visitor_category_end)
        ws.merge_cells(home_category_start + ":" + home_category_end)
        ws[visitor_category_start] = col[1]
        ws[visitor_category_start].fill = PatternFill(
            start_color=visitor_color, fill_type="solid"
        )
        ws[visitor_category_start].font = Font(color=colors.WHITE)
        ws[home_category_start] = col[1]
        ws[home_category_start].fill = PatternFill(
            start_color=home_color, fill_type="solid"
        )
        ws[home_category_start].font = Font(color=colors.WHITE)
        row += 1

        visitor_player_name = col_a + str(row)
        visitor_season_avg = col_b + str(row)
        visitor_last3 = col_c + str(row)
        visitor_lines = col_d + str(row)
        visitor_odds = col_e + str(row)
        home_player_name = col_f + str(row)
        home_season_avg = col_g + str(row)
        home_last3 = col_h + str(row)
        home_lines = col_i + str(row)
        home_odds = col_j + str(row)

        ws[visitor_player_name] = "Player Name"
        ws[visitor_season_avg] = "Season Avg"
        ws[visitor_last3] = "Last 3 Games"
        ws[visitor_lines] = "O/U Line"
        ws[visitor_odds] = "Odds O/U"

        ws[home_player_name] = "Player Name"
        ws[home_season_avg] = "Season Avg"
        ws[home_last3] = "Last 3 Games"
        ws[home_lines] = "O/U Line"
        ws[home_odds] = "Odds O/U"

        row += 1

        for index, player in visitor_player_stats[0:6].iterrows():
            id = int(player["Player_ID"])
            visitor_player = col_a + str(row)
            visitor_stat_avg = col_b + str(row)
            visitor_stat_last3 = col_c + str(row)
            visitor_line = col_d + str(row)
            vistor_odd = col_e + str(row)
            player_name: str = player["Player"]
            ws[visitor_player] = player_name
            for injury_index, injury in injuries.iterrows():
                if player["Player"] == injury["Player"]:
                    if injury["Status"] == "DTD":
                        ws[visitor_player] = player_name + '*'
            ws[visitor_stat_avg] = player[col[0]]
            ws[visitor_stat_last3] = get_recent_stats(id, visitor_player_logs, col[0])
            try:
                odds = Database('nba').select_player_odds(player=player_name, field=col[2])
                ws[visitor_line] = odds[0]
                ws[vistor_odd] = odds[1]
            except:
                ws[visitor_line] = "-"
                ws[vistor_odd] = "-"
            row += 1

        row -= 6

        for index, player in home_player_stats[0:6].iterrows():
            id = int(player["Player_ID"])
            home_player = col_f + str(row)
            home_stat_avg = col_g + str(row)
            home_stat_last3 = col_h + str(row)
            home_line = col_i + str(row)
            home_odd = col_j + str(row)
            player_name: str = player["Player"]
            ws[home_player] = player_name
            for injury_index, injury in injuries.iterrows():
                if player["Player"] == injury["Player"]:
                    if injury["Status"] == "DTD":
                        ws[home_player] = player_name + '*'
            ws[home_stat_avg] = player[col[0]]
            ws[home_stat_last3] = get_recent_stats(id, home_player_logs, col[0])
            try:
                odds = Database('nba').select_player_odds(player=player_name, field=col[2])
                ws[home_line] = odds[0]
                ws[home_odd] = odds[1]
            except:
                ws[home_line] = "-"
                ws[home_odd] = "-"

            row += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 20
    ws.column_dimensions["I"].width = 20
    ws.column_dimensions["J"].width = 20

    ws.merge_cells("A55:J55")
    ws[
        "A55"
    ] = "* Players marked with asterisks have been listed as Day To Day and could be questionable for the game"

    for col in ws.columns:
        for cell in col:
            alignment_obj = copy(cell.alignment)
            alignment_obj.horizontal = "center"
            alignment_obj.vertical = "center"
            cell.alignment = alignment_obj
            bold_obj = copy(cell.font)
            bold_obj.bold = True
            cell.font = bold_obj

    ws.merge_cells("E2:F6")
    pil_image = Image.open("assets/logo-transparent.png")
    pil_image = pil_image.resize((280, 100))
    temp_image_path = "assets/temp1.png"
    pil_image.save(temp_image_path)
    img = XLImage(temp_image_path)
    ws.add_image(img, "E2")

    ws.merge_cells("A1:B6")
    img = Image.open(
        "assets/logos/" + visitor_team_name.lower().replace(" ", "_") + ".png"
    )
    img.thumbnail((2240, 960), Image.LANCZOS)
    img_w, img_h = img.size
    background = Image.new("RGBA", (2800, 1200), (255, 255, 255, 255))
    bg_w, bg_h = background.size
    offset = ((bg_w - img_w) // 2, (bg_h - img_h) // 2)
    background.paste(img, offset)
    background.thumbnail((280, 120), Image.LANCZOS)
    temp_image_path = "assets/temp2.png"
    background.save(temp_image_path)
    img = XLImage(temp_image_path)
    ws.add_image(img, "A1")

    ws.merge_cells("I1:J6")
    img = Image.open(
        "./assets/logos/" + home_team_name.lower().replace(" ", "_") + ".png"
    )
    img.thumbnail((2240, 960), Image.LANCZOS)
    img_w, img_h = img.size
    background = Image.new("RGBA", (2800, 1200), (255, 255, 255, 255))
    bg_w, bg_h = background.size
    offset = ((bg_w - img_w) // 2, (bg_h - img_h) // 2)
    background.paste(img, offset)
    background.thumbnail((280, 120), Image.LANCZOS)
    temp_image_path = "./assets/temp3.png"
    background.save(temp_image_path)
    img = XLImage(temp_image_path)
    ws.add_image(img, "I1")

    home_abv = home["Abbreviation"].values[0]
    visitor_abv = visitor["Abbreviation"].values[0]
    file_name = str(visitor_abv) + "@" + str(home_abv) + "-players-" + date
    path = "./com/statlume/nba/sheets/" + file_name + ".xlsx"
    wb.save(path)
    os.remove("assets/temp1.png")
    os.remove("assets/temp2.png")
    os.remove("assets/temp3.png")
    load_image(path=path, file_name=file_name)


def create_txt_file(visitor: pd.DataFrame, home: pd.DataFrame) -> None:
    visitor: str = visitor["Team"].values[0]
    home: str = home["Team"].values[0]
    line = visitor + " at " + home
    with open("com/statlume/nba/games.txt", "a") as file:
        file.write(line + '\n')
        file.close()


def create_sheets(
    date: str,
    home: pd.DataFrame,
    visitor: pd.DataFrame,
    home_player_stats: pd.DataFrame,
    visitor_player_stats: pd.DataFrame,
    home_player_logs: pd.DataFrame,
    visitor_player_logs: pd.DataFrame,
):
    create_team_sheet(date=date, home=home, visitor=visitor)
    create_player_sheet(
        date=date,
        home=home,
        visitor=visitor,
        home_player_stats=home_player_stats,
        visitor_player_stats=visitor_player_stats,
        home_player_logs=home_player_logs,
        visitor_player_logs=visitor_player_logs,
    )
    create_txt_file(home=home, visitor=visitor)


def load_image(path: str, file_name: str):
    image_file = "./com/statlume/nba/images/" + file_name + ".png"
    excel2img.export_img(path, image_file, "Sheet", None)


def delete_matchups():
    file = "./com/statlume/nba/games.txt"
    if os.path.isfile(file):
        os.remove(file)

    images = "./com/statlume/nba/images/"
    for filename in os.listdir(images):
        file_path = os.path.join(images, filename)
        if os.path.isfile(file_path):
            os.remove(file_path)

    sheets = "./com/statlume/nba/sheets/"
    for filename in os.listdir(sheets):
        file_path = os.path.join(sheets, filename)
        if os.path.isfile(file_path):
            os.remove(file_path)


def get_todays_games() -> pd.DataFrame:
    today = datetime.datetime.today()
    yesterday = today - datetime.timedelta(days=1)
    sql = (
        "SELECT * FROM schedule WHERE schedule.GameDate > '%s' and schedule.GameDate < '%s';"
        % (yesterday, today)
    )
    todays_games = Database("nba").select_table_as_df("schedule", sql=sql)
    return todays_games


def get_matchups():
    games = get_todays_games()
    teamstats = Database("nba").select_table_as_df("teamstats")
    playerstats = Database("nba").select_table_as_df("playerstats")
    gamelogs = Database("nba").select_table_as_df("gamelogs")
    injuries = Database("nba").select_table_as_df("injuries").drop_duplicates()

    for index, row in games.iterrows():
        date = row["GameDate"]
        date = str(date.strftime("%m-%d-%Y"))
        home_team: str = row["Home"]
        visitor_team: str = row["Visitor"]
        home = teamstats[teamstats["Team"] == home_team]
        visitor = teamstats[teamstats["Team"] == visitor_team]
        home_player_stats: pd.DataFrame = playerstats[playerstats["Team"] == home_team]
        visitor_player_stats: pd.DataFrame = playerstats[
            playerstats["Team"] == visitor_team
        ]

        for player_index, player in visitor_player_stats.iterrows():
            for injury_index, injury in injuries.iterrows():
                if player["Player"] == injury["Player"]:
                    if injury["Status"] == "O":
                        visitor_player_stats = visitor_player_stats.drop(labels=player_index)


        for player_index, player in home_player_stats.iterrows():
            for injury_index, injury in injuries.iterrows():
                if player["Player"] == injury["Player"]:
                    if injury["Status"] == "O":
                        home_player_stats = home_player_stats.drop(labels=player_index)

        visitor_player_logs = pd.DataFrame()
        for index, player in visitor_player_stats[0:6].iterrows():
            id = int(player["Player_ID"])
            visitor_player_logs = pd.concat(
                [visitor_player_logs, gamelogs[gamelogs["Player_ID"] == id]]
            )

        home_player_logs = pd.DataFrame()
        for index, player in home_player_stats[0:6].iterrows():
            id = int(player["Player_ID"])
            home_player_logs = pd.concat(
                [home_player_logs, gamelogs[gamelogs["Player_ID"] == id]]
            )

        create_sheets(
            date=date,
            home=home,
            visitor=visitor,
            home_player_stats=home_player_stats,
            visitor_player_stats=visitor_player_stats,
            home_player_logs=home_player_logs,
            visitor_player_logs=visitor_player_logs,
        )


if __name__ == "__main__":
    delete_matchups()
    get_matchups()
